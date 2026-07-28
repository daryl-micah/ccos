"""Parse uploaded CSV / Excel files into influencer rows.

Lets teams bulk-load their existing creator spreadsheet into CCOS,
removing the migration barrier (PRODUCT.md integration phase 1).
"""

import csv
import io
import re
from urllib.parse import urlparse

from openpyxl import load_workbook

# Map common spreadsheet headers (lowercased) to Influencer fields.
FIELD_ALIASES = {
    "name": "name",
    "creator": "name",
    "influencer": "name",
    "instagram": "instagram_username",
    "instagram_username": "instagram_username",
    "instagram link": "instagram_username",
    "instagram_link": "instagram_username",
    "handle": "instagram_username",
    "ig": "instagram_username",
    "youtube": "youtube_channel",
    "youtube_channel": "youtube_channel",
    "city": "city",
    "country": "country",
    "category": "category",
    "language": "language",
    "manager": "manager_name",
    "manager_name": "manager_name",
    "email": "email",
    "phone": "phone",
    "contact": "phone",
    "contact details": "phone",
    "contact number": "phone",
    "contact_number": "phone",
    "notes": "notes",
    # Generic profile-link columns: the URL is parsed to detect the
    # platform, rather than assumed to be a specific one.
    "link": "_profile_link",
    "links": "_profile_link",
    "profile link": "_profile_link",
    "profile_link": "_profile_link",
    "profile url": "_profile_link",
    "profile_url": "_profile_link",
    "url": "_profile_link",
    "social link": "_profile_link",
}

ALLOWED_FIELDS = {f for f in FIELD_ALIASES.values() if not f.startswith("_")}

_IG_SKIP_SEGMENTS = {"p", "reel", "reels", "stories", "explore", "tv"}
_YT_PATH_SEGMENTS = {"channel", "c", "user"}


def _parse_profile_url(value: str) -> tuple[str, str] | None:
    """Detect Instagram/YouTube profile links and extract (field, handle)."""
    text = value.strip()
    if not text:
        return None
    candidate = text if re.match(r"^https?://", text, re.I) else f"https://{text}"
    try:
        parsed = urlparse(candidate)
    except ValueError:
        return None
    host = parsed.netloc.lower().removeprefix("www.")
    segments = [s for s in parsed.path.split("/") if s]
    if not segments:
        return None
    if "instagram.com" in host:
        handle = segments[0].lstrip("@")
        if handle.lower() in _IG_SKIP_SEGMENTS:
            return None
        return ("instagram_username", handle)
    if "youtube.com" in host:
        if segments[0].startswith("@"):
            return ("youtube_channel", segments[0])
        if segments[0] in _YT_PATH_SEGMENTS and len(segments) > 1:
            return ("youtube_channel", segments[1])
        return None
    return None


def _normalize_row(raw: dict[str, object]) -> dict[str, str]:
    mapped: dict[str, str] = {}
    for header, value in raw.items():
        if header is None:
            continue
        field = FIELD_ALIASES.get(str(header).strip().lower())
        if not field:
            continue
        text = "" if value is None else str(value).strip()
        if not text:
            continue
        if field == "_profile_link":
            parsed = _parse_profile_url(text)
            if parsed:
                mapped.setdefault(*parsed)
            continue
        mapped[field] = text
    # A discrete instagram/youtube column may itself hold a full URL
    # (e.g. an "instagram link" header) - normalize it to a bare handle.
    for field in ("instagram_username", "youtube_channel"):
        if field in mapped:
            parsed = _parse_profile_url(mapped[field])
            if parsed:
                if parsed[0] != field:
                    del mapped[field]
                mapped[parsed[0]] = parsed[1]
    return mapped


def parse_influencer_links(text: str) -> tuple[list[dict[str, str]], list[str]]:
    """Parse a pasted list of profile URLs (one per line, or comma-separated).

    Returns (rows, skipped) where each row is a minimal influencer dict
    (name defaults to the handle) and skipped holds tokens that weren't
    recognized as an Instagram/YouTube profile link.
    """
    rows: list[dict[str, str]] = []
    skipped: list[str] = []
    seen: set[str] = set()
    tokens = [t.strip() for line in text.splitlines() for t in line.split(",")]
    for token in tokens:
        if not token:
            continue
        parsed = _parse_profile_url(token)
        if not parsed:
            skipped.append(token)
            continue
        field, handle = parsed
        key = f"{field}:{handle.lower()}"
        if key in seen:
            continue
        seen.add(key)
        rows.append({"name": handle, field: handle})
    return rows, skipped


def parse_influencer_rows(filename: str, content: bytes) -> list[dict[str, str]]:
    """Return normalized influencer dicts from a CSV or XLSX upload."""
    lower = filename.lower()
    if lower.endswith(".csv"):
        rows = _parse_csv(content)
    elif lower.endswith((".xlsx", ".xlsm")):
        rows = _parse_xlsx(content)
    else:
        raise ValueError("Unsupported file type. Upload a .csv or .xlsx file.")

    normalized = [_normalize_row(r) for r in rows]
    return [r for r in normalized if r.get("name")]


def _parse_csv(content: bytes) -> list[dict[str, object]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def _parse_xlsx(content: bytes) -> list[dict[str, object]]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
    except StopIteration:
        return []
    out: list[dict[str, object]] = []
    for row in rows_iter:
        if row is None or all(c is None for c in row):
            continue
        out.append(dict(zip(headers, row, strict=False)))
    wb.close()
    return out
