"""Excel report generation for a campaign.

Produces a readable workbook mapped to the columns marketing teams
already use, so adopting CCOS doesn't force abandoning their spreadsheet
workflow (see REQUIREMENT_DOC "Reports").
"""

import io
import uuid
from collections import defaultdict
from dataclasses import dataclass
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import (
    Agency,
    Campaign,
    CampaignInfluencer,
    Deliverable,
    Influencer,
    Metric,
    Post,
)

# Insight metrics surfaced as columns in the summary / posts sheets.
KEY_METRICS = [
    "likes",
    "comments",
    "views",
    "engagement_rate",
    "engagement_rate_reach",
]

HEADER_FONT = Font(bold=True)


def _num(value: Decimal | None) -> float | None:
    return float(value) if value is not None else None


def _write_header(ws: Worksheet, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT


def _autosize(ws: Worksheet) -> None:
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            width + 2, 60
        )


# Derived per-unit metrics that must be averaged, never summed, across rows.
_RATE_METRICS = {"cpv", "cpm", "cpa", "roas"}


def _aggregate(values: list[float], metric_name: str) -> float:
    """Rates/derived metrics average; counts sum."""
    if not values:
        return 0.0
    if (
        metric_name.endswith("_rate")
        or metric_name.endswith("_rate_reach")
        or metric_name in _RATE_METRICS
    ):
        return round(sum(values) / len(values), 4)
    return round(sum(values), 4)


def _slug(name: str) -> str:
    return "".join(c if c.isalnum() else "_" for c in name).strip("_") or "campaign"


def _xlsx(wb: Workbook, filename: str) -> tuple[io.BytesIO, str]:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, filename


async def build_campaign_report(
    db: AsyncSession, campaign_id: uuid.UUID
) -> tuple[io.BytesIO, str] | None:
    """Return (xlsx bytes, filename) for the campaign, or None if not found."""
    campaign = await db.scalar(
        select(Campaign).where(
            Campaign.id == campaign_id, Campaign.deleted_at.is_(None)
        )
    )
    if campaign is None:
        return None

    cis = list(
        await db.scalars(
            select(CampaignInfluencer).where(
                CampaignInfluencer.campaign_id == campaign_id,
                CampaignInfluencer.deleted_at.is_(None),
            )
        )
    )
    ci_ids = [ci.id for ci in cis]
    inf_ids = [ci.influencer_id for ci in cis]

    influencers = {
        i.id: i
        for i in await db.scalars(
            select(Influencer).where(Influencer.id.in_(inf_ids or [uuid.uuid4()]))
        )
    }
    deliverables = list(
        await db.scalars(
            select(Deliverable).where(
                Deliverable.campaign_influencer_id.in_(ci_ids or [uuid.uuid4()]),
                Deliverable.deleted_at.is_(None),
            )
        )
    )
    posts = list(
        await db.scalars(
            select(Post).where(
                Post.campaign_influencer_id.in_(ci_ids or [uuid.uuid4()]),
                Post.deleted_at.is_(None),
            )
        )
    )
    metrics = list(
        await db.scalars(
            select(Metric).where(
                Metric.campaign_influencer_id.in_(ci_ids or [uuid.uuid4()]),
                Metric.deleted_at.is_(None),
            )
        )
    )

    # Index helpers.
    deliv_by_ci: dict[uuid.UUID, list[Deliverable]] = defaultdict(list)
    for d in deliverables:
        deliv_by_ci[d.campaign_influencer_id].append(d)

    posts_by_ci: dict[uuid.UUID, list[Post]] = defaultdict(list)
    for p in posts:
        posts_by_ci[p.campaign_influencer_id].append(p)

    metrics_by_ci: dict[uuid.UUID, list[Metric]] = defaultdict(list)
    metrics_by_post: dict[uuid.UUID, list[Metric]] = defaultdict(list)
    for m in metrics:
        metrics_by_ci[m.campaign_influencer_id].append(m)
        if m.post_id:
            metrics_by_post[m.post_id].append(m)

    wb = Workbook()
    _build_summary_sheet(wb, campaign, cis, influencers, deliv_by_ci, posts_by_ci, metrics_by_ci)
    _build_deliverables_sheet(wb, cis, influencers, deliv_by_ci)
    _build_posts_sheet(wb, cis, influencers, posts_by_ci, metrics_by_post)
    _build_metrics_sheet(wb, cis, influencers, metrics, posts)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    slug = "".join(c if c.isalnum() else "_" for c in campaign.name).strip("_")
    return buf, f"{slug or 'campaign'}_report.xlsx"


def _build_summary_sheet(wb, campaign, cis, influencers, deliv_by_ci, posts_by_ci, metrics_by_ci):
    ws = wb.active
    ws.title = "Summary"

    # Campaign header block.
    ws.append(["Campaign", campaign.name])
    ws.append(["Brand", campaign.brand or ""])
    ws.append(["Objective", campaign.objective or ""])
    ws.append(["Status", campaign.status])
    ws.append(["Budget", _num(campaign.budget)])
    ws.append(
        ["Dates", f"{campaign.start_date or '—'} → {campaign.end_date or '—'}"]
    )
    for row in ws.iter_rows(min_row=1, max_row=6, max_col=1):
        row[0].font = HEADER_FONT
    ws.append([])

    header_row = ws.max_row + 1
    headers = ["Creator", "City", "Category", "Status", "Cost", "Deliverables", "Posts"]
    headers += [m.replace("_", " ").title() for m in KEY_METRICS]
    ws.append(headers)
    for cell in ws[header_row]:
        cell.font = HEADER_FONT

    for ci in cis:
        inf = influencers.get(ci.influencer_id)
        ci_metrics = metrics_by_ci.get(ci.id, [])
        by_name: dict[str, list[float]] = defaultdict(list)
        for m in ci_metrics:
            by_name[m.metric_name].append(float(m.metric_value))
        row = [
            inf.name if inf else "Unknown",
            inf.city if inf else "",
            inf.category if inf else "",
            ci.status,
            _num(ci.cost),
            len(deliv_by_ci.get(ci.id, [])),
            len(posts_by_ci.get(ci.id, [])),
        ]
        row += [_aggregate(by_name.get(m, []), m) for m in KEY_METRICS]
        ws.append(row)

    _autosize(ws)


def _build_deliverables_sheet(wb, cis, influencers, deliv_by_ci):
    ws = wb.create_sheet("Deliverables")
    _write_header(
        ws, ["Creator", "Type", "Quantity", "Due", "Posted", "Status", "Link"]
    )
    ci_by_id = {ci.id: ci for ci in cis}
    for ci_id, items in deliv_by_ci.items():
        inf = influencers.get(ci_by_id[ci_id].influencer_id) if ci_id in ci_by_id else None
        for d in items:
            ws.append(
                [
                    inf.name if inf else "Unknown",
                    d.type,
                    d.quantity,
                    str(d.due_date) if d.due_date else "",
                    str(d.posted_date) if d.posted_date else "",
                    d.status,
                    d.link or "",
                ]
            )
    _autosize(ws)


def _build_posts_sheet(wb, cis, influencers, posts_by_ci, metrics_by_post):
    ws = wb.create_sheet("Posts")
    # Collect all metric names present on posts (key metrics first).
    extra = sorted(
        {
            m.metric_name
            for ms in metrics_by_post.values()
            for m in ms
            if m.metric_name not in KEY_METRICS
        }
    )
    metric_cols = KEY_METRICS + extra
    _write_header(
        ws,
        ["Creator", "Platform", "URL", "Posted at"]
        + [m.replace("_", " ").title() for m in metric_cols],
    )
    ci_by_id = {ci.id: ci for ci in cis}
    for ci_id, items in posts_by_ci.items():
        inf = influencers.get(ci_by_id[ci_id].influencer_id) if ci_id in ci_by_id else None
        for p in items:
            latest: dict[str, float] = {}
            for m in metrics_by_post.get(p.id, []):
                latest[m.metric_name] = float(m.metric_value)
            ws.append(
                [
                    inf.name if inf else "Unknown",
                    p.platform,
                    p.url,
                    str(p.posted_at) if p.posted_at else "",
                ]
                + [latest.get(m, "") for m in metric_cols]
            )
    _autosize(ws)


def _build_metrics_sheet(wb, cis, influencers, metrics, posts):
    ws = wb.create_sheet("Metrics")
    _write_header(
        ws, ["Creator", "Post URL", "Metric", "Value", "Source", "Captured at"]
    )
    ci_by_id = {ci.id: ci for ci in cis}
    post_url = {p.id: p.url for p in posts}
    for m in metrics:
        ci = ci_by_id.get(m.campaign_influencer_id)
        inf = influencers.get(ci.influencer_id) if ci else None
        ws.append(
            [
                inf.name if inf else "Unknown",
                post_url.get(m.post_id, "") if m.post_id else "",
                m.metric_name,
                float(m.metric_value),
                m.source,
                str(m.captured_at) if m.captured_at else "",
            ]
        )
    _autosize(ws)


# --- Focused exports (creators / posts / overall tracker) --------------------


@dataclass
class _Bundle:
    campaign: Campaign
    cis: list[CampaignInfluencer]
    influencers: dict[uuid.UUID, Influencer]
    deliv_by_ci: dict[uuid.UUID, list[Deliverable]]
    posts_by_ci: dict[uuid.UUID, list[Post]]
    metrics_by_ci: dict[uuid.UUID, list[Metric]]
    metrics_by_post: dict[uuid.UUID, list[Metric]]
    followers: dict[uuid.UUID, float]  # influencer_id -> latest followers
    repeat: set[uuid.UUID]  # influencers worked with on >1 campaign
    agencies: dict[uuid.UUID, Agency]  # agency_id -> Agency ("closed by")


async def _latest_followers(db: AsyncSession, inf_ids: list[uuid.UUID]) -> dict:
    rows = await db.scalars(
        select(Metric)
        .where(
            Metric.influencer_id.in_(inf_ids or [uuid.uuid4()]),
            Metric.metric_name == "followers",
            Metric.deleted_at.is_(None),
        )
        .order_by(Metric.captured_at)  # ascending → last write wins (latest)
    )
    out: dict[uuid.UUID, float] = {}
    for m in rows:
        out[m.influencer_id] = float(m.metric_value)
    return out


async def _repeat_influencers(db: AsyncSession, inf_ids: list[uuid.UUID]) -> set:
    rows = await db.scalars(
        select(CampaignInfluencer).where(
            CampaignInfluencer.influencer_id.in_(inf_ids or [uuid.uuid4()]),
            CampaignInfluencer.deleted_at.is_(None),
        )
    )
    campaigns_per_inf: dict[uuid.UUID, set] = defaultdict(set)
    for ci in rows:
        campaigns_per_inf[ci.influencer_id].add(ci.campaign_id)
    return {inf for inf, camps in campaigns_per_inf.items() if len(camps) > 1}


async def _load_bundle(db: AsyncSession, campaign_id: uuid.UUID) -> _Bundle | None:
    campaign = await db.scalar(
        select(Campaign).where(
            Campaign.id == campaign_id, Campaign.deleted_at.is_(None)
        )
    )
    if campaign is None:
        return None

    cis = list(
        await db.scalars(
            select(CampaignInfluencer).where(
                CampaignInfluencer.campaign_id == campaign_id,
                CampaignInfluencer.deleted_at.is_(None),
            )
        )
    )
    ci_ids = [ci.id for ci in cis]
    inf_ids = [ci.influencer_id for ci in cis]

    influencers = {
        i.id: i
        for i in await db.scalars(
            select(Influencer).where(Influencer.id.in_(inf_ids or [uuid.uuid4()]))
        )
    }
    deliverables = await db.scalars(
        select(Deliverable).where(
            Deliverable.campaign_influencer_id.in_(ci_ids or [uuid.uuid4()]),
            Deliverable.deleted_at.is_(None),
        )
    )
    posts = await db.scalars(
        select(Post).where(
            Post.campaign_influencer_id.in_(ci_ids or [uuid.uuid4()]),
            Post.deleted_at.is_(None),
        )
    )
    metrics = await db.scalars(
        select(Metric).where(
            Metric.campaign_influencer_id.in_(ci_ids or [uuid.uuid4()]),
            Metric.deleted_at.is_(None),
        )
    )

    deliv_by_ci: dict[uuid.UUID, list] = defaultdict(list)
    for d in deliverables:
        deliv_by_ci[d.campaign_influencer_id].append(d)
    posts_by_ci: dict[uuid.UUID, list] = defaultdict(list)
    for p in posts:
        posts_by_ci[p.campaign_influencer_id].append(p)
    metrics_by_ci: dict[uuid.UUID, list] = defaultdict(list)
    metrics_by_post: dict[uuid.UUID, list] = defaultdict(list)
    for m in metrics:
        metrics_by_ci[m.campaign_influencer_id].append(m)
        if m.post_id:
            metrics_by_post[m.post_id].append(m)

    agency_ids = [ci.agency_id for ci in cis if ci.agency_id]
    agencies = {
        a.id: a
        for a in await db.scalars(
            select(Agency).where(Agency.id.in_(agency_ids or [uuid.uuid4()]))
        )
    }

    return _Bundle(
        campaign=campaign,
        cis=cis,
        influencers=influencers,
        deliv_by_ci=deliv_by_ci,
        posts_by_ci=posts_by_ci,
        metrics_by_ci=metrics_by_ci,
        metrics_by_post=metrics_by_post,
        followers=await _latest_followers(db, inf_ids),
        repeat=await _repeat_influencers(db, inf_ids),
        agencies=agencies,
    )


def _metric_columns(metric_lists, exclude: set[str] | None = None) -> list[str]:
    """KEY_METRICS first, then any other metric names present, sorted."""
    exclude = exclude or set()
    extra = sorted(
        {
            m.metric_name
            for ms in metric_lists
            for m in ms
            if m.metric_name not in KEY_METRICS and m.metric_name not in exclude
        }
    )
    return [m for m in KEY_METRICS if m not in exclude] + extra


async def build_campaign_creators_report(
    db: AsyncSession, campaign_id: uuid.UUID
) -> tuple[io.BytesIO, str] | None:
    """Campaign-wise creator sheet: one row per creator with aggregated metrics."""
    b = await _load_bundle(db, campaign_id)
    if b is None:
        return None

    metric_cols = _metric_columns(b.metrics_by_ci.values())
    wb = Workbook()
    ws = wb.active
    ws.title = "Creators"
    _write_header(
        ws,
        [
            "Creator",
            "Instagram",
            "Closed by",
            "Contact",
            "City",
            "Category",
            "Status",
            "Cost",
            "Worked before",
            "Deliverables",
            "Posts",
            "Followers",
        ]
        + [m.replace("_", " ").title() for m in metric_cols]
        + ["Remarks"],
    )

    for ci in b.cis:
        inf = b.influencers.get(ci.influencer_id)
        agency = b.agencies.get(ci.agency_id) if ci.agency_id else None
        by_name: dict[str, list[float]] = defaultdict(list)
        for m in b.metrics_by_ci.get(ci.id, []):
            by_name[m.metric_name].append(float(m.metric_value))
        row = [
            inf.name if inf else "Unknown",
            (inf.instagram_username if inf else None) or "",
            agency.name if agency else "In-house",
            ((inf.email or inf.phone) if inf else None) or "",
            (inf.city if inf else None) or "",
            (inf.category if inf else None) or "",
            ci.status,
            _num(ci.cost),
            "Yes" if inf and inf.id in b.repeat else "No",
            len(b.deliv_by_ci.get(ci.id, [])),
            len(b.posts_by_ci.get(ci.id, [])),
            b.followers.get(ci.influencer_id, ""),
        ]
        row += [
            _aggregate(by_name[m], m) if m in by_name else "" for m in metric_cols
        ]
        row.append(ci.remarks or "")
        ws.append(row)

    _autosize(ws)
    return _xlsx(wb, f"{_slug(b.campaign.name)}_creators.xlsx")


async def build_campaign_posts_report(
    db: AsyncSession, campaign_id: uuid.UUID
) -> tuple[io.BytesIO, str] | None:
    """Campaign-wise posts sheet: one row per live post with its metrics."""
    b = await _load_bundle(db, campaign_id)
    if b is None:
        return None

    metric_cols = _metric_columns(b.metrics_by_post.values())
    deliv_by_id = {d.id: d for ds in b.deliv_by_ci.values() for d in ds}
    ci_by_id = {ci.id: ci for ci in b.cis}

    wb = Workbook()
    ws = wb.active
    ws.title = "Posts"
    _write_header(
        ws,
        ["Posted at", "Creator", "Instagram", "Platform", "Live link", "Deliverable"]
        + [m.replace("_", " ").title() for m in metric_cols],
    )
    for ci_id, items in b.posts_by_ci.items():
        ci = ci_by_id.get(ci_id)
        inf = b.influencers.get(ci.influencer_id) if ci else None
        for p in items:
            latest: dict[str, float] = {}
            for m in b.metrics_by_post.get(p.id, []):
                latest[m.metric_name] = float(m.metric_value)
            deliv = deliv_by_id.get(p.deliverable_id) if p.deliverable_id else None
            ws.append(
                [
                    str(p.posted_at) if p.posted_at else "",
                    inf.name if inf else "Unknown",
                    (inf.instagram_username if inf else None) or "",
                    p.platform,
                    p.url,
                    deliv.type if deliv else "",
                ]
                + [latest.get(m, "") for m in metric_cols]
            )

    _autosize(ws)
    return _xlsx(wb, f"{_slug(b.campaign.name)}_posts.xlsx")


# Single-sheet "POA - Supply" export — mirrors the marketing team's master
# tracker (one row per live post). Only metrics that actually populate are
# surfaced; untracked columns (payment status, shares, performance cuts,
# leads) are kept as blank placeholders to preserve the familiar layout.
POA_HEADERS = [
    "Live Month",
    "Platform",
    "Content Type",
    "Agency Name",
    "Creator Name",
    "Profile",
    "City",
    "Language",
    "Amount Paid",
    "Live Status",
    "Payment status",
    "Video Live Link",
    "Views",
    "Comments",
    "Likes",
    "Shares",
    "ER %",
    "Actual CPV",
    "Performance cuts",
    "Leads Generated",
]


async def build_campaign_poa_report(
    db: AsyncSession, campaign_id: uuid.UUID
) -> tuple[io.BytesIO, str] | None:
    """Single 'POA - Supply' sheet: one row per live post (master-tracker layout)."""
    b = await _load_bundle(db, campaign_id)
    if b is None:
        return None

    deliv_by_id = {d.id: d for ds in b.deliv_by_ci.values() for d in ds}

    wb = Workbook()
    ws = wb.active
    ws.title = "POA - Supply"
    _write_header(ws, POA_HEADERS)

    for ci in b.cis:
        inf = b.influencers.get(ci.influencer_id)
        agency = b.agencies.get(ci.agency_id) if ci.agency_id else None
        agency_name = agency.name if agency else "In-house"
        creator = inf.name if inf else "Unknown"
        profile = (
            f"https://www.instagram.com/{inf.instagram_username}"
            if inf and inf.instagram_username
            else ""
        )
        city = (inf.city if inf else None) or ""
        language = (inf.language if inf else None) or ""
        cost = _num(ci.cost)
        category = (inf.category if inf else None) or ""

        posts = b.posts_by_ci.get(ci.id, [])
        if not posts:
            # No live post yet — emit a WIP placeholder row.
            ws.append(
                [
                    "",
                    "Instagram",
                    category,
                    agency_name,
                    creator,
                    profile,
                    city,
                    language,
                    cost,
                    ci.status.replace("_", " ").title(),
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                ]
            )
            continue

        for p in posts:
            latest: dict[str, float] = {}
            for m in b.metrics_by_post.get(p.id, []):
                latest[m.metric_name] = float(m.metric_value)
            views = latest.get("views")
            deliv = deliv_by_id.get(p.deliverable_id) if p.deliverable_id else None
            content_type = (deliv.type if deliv else None) or category
            cpv = (
                round(cost / views, 4) if cost is not None and views else ""
            )
            ws.append(
                [
                    str(p.posted_at.date()) if p.posted_at else "",
                    p.platform or "Instagram",
                    content_type,
                    agency_name,
                    creator,
                    profile,
                    city,
                    language,
                    cost,
                    "Live",
                    "",  # Payment status — not tracked
                    p.url,
                    views if views is not None else "",
                    latest.get("comments", ""),
                    latest.get("likes", ""),
                    "",  # Shares — not tracked
                    latest.get("engagement_rate", ""),
                    cpv,
                    "",  # Performance cuts — not tracked
                    "",  # Leads Generated — not tracked
                ]
            )

    _autosize(ws)
    return _xlsx(wb, f"{_slug(b.campaign.name)}_poa.xlsx")


async def build_tracker_report(db: AsyncSession) -> tuple[io.BytesIO, str]:
    """Overall campaigns tracker: one row per campaign with aggregated funnel."""
    campaigns = list(
        await db.scalars(
            select(Campaign)
            .where(Campaign.deleted_at.is_(None))
            .order_by(Campaign.created_at)
        )
    )
    cis = list(
        await db.scalars(
            select(CampaignInfluencer).where(CampaignInfluencer.deleted_at.is_(None))
        )
    )
    posts = list(
        await db.scalars(select(Post).where(Post.deleted_at.is_(None)))
    )
    metrics = list(
        await db.scalars(
            select(Metric).where(
                Metric.deleted_at.is_(None),
                Metric.campaign_influencer_id.is_not(None),
            )
        )
    )

    ci_to_campaign = {ci.id: ci.campaign_id for ci in cis}
    cis_by_campaign: dict[uuid.UUID, list] = defaultdict(list)
    for ci in cis:
        cis_by_campaign[ci.campaign_id].append(ci)
    posts_by_campaign: dict[uuid.UUID, int] = defaultdict(int)
    for p in posts:
        cid = ci_to_campaign.get(p.campaign_influencer_id)
        if cid:
            posts_by_campaign[cid] += 1

    # campaign_id -> {metric_name -> [values]}; post metrics roll up too.
    by_campaign: dict[uuid.UUID, dict[str, list]] = defaultdict(
        lambda: defaultdict(list)
    )
    for m in metrics:
        cid = ci_to_campaign.get(m.campaign_influencer_id)
        if cid:
            by_campaign[cid][m.metric_name].append(float(m.metric_value))

    # revenue/roas get explicit columns below; don't duplicate them.
    extra = sorted(
        {
            name
            for camp in by_campaign.values()
            for name in camp
            if name not in KEY_METRICS and name not in {"revenue", "roas"}
        }
    )
    metric_cols = KEY_METRICS + extra

    wb = Workbook()
    ws = wb.active
    ws.title = "Tracker"
    _write_header(
        ws,
        ["Campaign", "Brand", "Status", "Budget", "Start", "End", "Creators", "Posts", "Spend"]
        + [m.replace("_", " ").title() for m in metric_cols]
        + ["Revenue", "ROAS"],
    )
    for c in campaigns:
        camp_cis = cis_by_campaign.get(c.id, [])
        spend = sum(float(ci.cost) for ci in camp_cis if ci.cost is not None)
        by_name = by_campaign.get(c.id, {})
        revenue = sum(by_name.get("revenue", []))
        roas = round(revenue / spend, 4) if spend > 0 and revenue else ""
        row = [
            c.name,
            c.brand or "",
            c.status,
            _num(c.budget),
            str(c.start_date) if c.start_date else "",
            str(c.end_date) if c.end_date else "",
            len(camp_cis),
            posts_by_campaign.get(c.id, 0),
            round(spend, 2),
        ]
        row += [
            _aggregate(by_name[m], m) if m in by_name else "" for m in metric_cols
        ]
        row += [round(revenue, 2) if revenue else "", roas]
        ws.append(row)

    _autosize(ws)
    return _xlsx(wb, "campaigns_tracker.xlsx")
